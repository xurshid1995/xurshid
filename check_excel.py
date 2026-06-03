# -*- coding: utf-8 -*-
"""Excel varaqlaridagi ma'lumotlarni tekshirish"""
import os
import sys
sys.path.insert(0, '/var/www/xurshid')
from dotenv import load_dotenv
load_dotenv('/var/www/xurshid/.env')
import psycopg2
from psycopg2.extras import RealDictCursor

c = psycopg2.connect(
    host=os.getenv('DB_HOST', 'localhost'),
    port=int(os.getenv('DB_PORT', 5432)),
    database=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
)
cur = c.cursor(cursor_factory=RealDictCursor)

print("=" * 60)

# 1. Mahsulotlar
cur.execute("SELECT COUNT(*) as cnt FROM products")
r = cur.fetchone()
print(f"📦 Mahsulotlar: {r['cnt']} ta")
cur.execute("SELECT id, name, cost_price, sell_price, unit_type FROM products LIMIT 2")
for row in cur.fetchall():
    print(f"   {dict(row)}")

print()

# 2. Mijozlar
cur.execute("SELECT COUNT(*) as cnt FROM customers")
r = cur.fetchone()
print(f"👥 Mijozlar: {r['cnt']} ta")
cur.execute("""
    SELECT c.id, c.name, c.phone, c.balance,
           COALESCE((SELECT ROUND(SUM(debt_usd)::numeric,2) FROM sales
                     WHERE customer_id=c.id AND payment_status='partial'),0) as qarz
    FROM customers c LIMIT 2
""")
for row in cur.fetchall():
    print(f"   {dict(row)}")

print()

# 3. Qarzlar
cur.execute("""
    SELECT COUNT(*) as cnt, ROUND(SUM(debt_usd)::numeric,2) as jami_usd,
           ROUND(SUM(debt_amount)::numeric,0) as jami_uzs
    FROM sales WHERE payment_status='partial' AND debt_usd > 0
""")
r = cur.fetchone()
print(f"💰 Aktiv qarzlar: {r['cnt']} ta | Jami: ${r['jami_usd']} / {r['jami_uzs']} UZS")
cur.execute("""
    SELECT TO_CHAR(s.sale_date,'YYYY-MM-DD') as sana,
           COALESCE(c.name,'?') as mijoz, s.debt_usd, s.debt_amount
    FROM sales s LEFT JOIN customers c ON c.id=s.customer_id
    WHERE s.payment_status='partial' AND s.debt_usd>0 LIMIT 2
""")
for row in cur.fetchall():
    print(f"   {dict(row)}")

print()

# 4. Sotuvlar 30 kun
cur.execute("""
    SELECT COUNT(*) as cnt,
           ROUND(SUM(total_amount)::numeric,2) as jami,
           ROUND(SUM(total_profit)::numeric,2) as foyda
    FROM sales WHERE sale_date >= NOW() - INTERVAL '30 days'
""")
r = cur.fetchone()
print(f"🛒 Sotuvlar (30 kun): {r['cnt']} ta | Jami: ${r['jami']} | Foyda: ${r['foyda']}")

print()

# 5. Xarajatlar 30 kun
cur.execute("""
    SELECT COUNT(*) as cnt,
           ROUND(SUM(amount_usd)::numeric,2) as jami_usd,
           ROUND(SUM(amount_uzs)::numeric,0) as jami_uzs
    FROM expenses WHERE expense_date >= NOW() - INTERVAL '30 days'
""")
r = cur.fetchone()
print(f"💸 Xarajatlar (30 kun): {r['cnt']} ta | ${r['jami_usd']} / {r['jami_uzs']} UZS")

print()
print("=" * 60)

# Excel faylni tekshirish
import openpyxl
import glob
files = sorted(glob.glob('/var/backups/xurshid/hisobot_*.xlsx'))
if files:
    latest = files[-1]
    wb = openpyxl.load_workbook(latest)
    print(f"Excel fayl: {latest}")
    for ws in wb.worksheets:
        rows = ws.max_row - 1  # header minus
        print(f"  Varaq '{ws.title}': {ws.max_row-1} qator, {ws.max_column} ustun")
        # Sarlavhalar
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        print(f"    Ustunlar: {headers}")

c.close()
