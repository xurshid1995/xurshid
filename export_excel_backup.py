# -*- coding: utf-8 -*-
"""
Kunlik backup uchun Excel hisobot generatori
Chaqirilishi: python export_excel_backup.py
Natija: /var/backups/xurshid/hisobot_YYYYMMDD_HHMMSS.xlsx
"""
import os
import sys
import re
from datetime import datetime

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Env yuklash
sys.path.insert(0, '/var/www/xurshid')
try:
    from dotenv import load_dotenv
    load_dotenv('/var/www/xurshid/.env')
except Exception:
    pass

# Avval alohida env o'zgaruvchilardan, keyin DATABASE_URL dan
if os.getenv('DB_NAME'):
    DB_CONFIG = {
        'host':     os.getenv('DB_HOST', 'localhost'),
        'port':     int(os.getenv('DB_PORT', 5432)),
        'database': os.getenv('DB_NAME', 'xurshid_db'),
        'user':     os.getenv('DB_USER', 'postgres'),
        'password': os.getenv('DB_PASSWORD', ''),
    }
else:
    DB_URL = os.getenv('DATABASE_URL', '')
    m = re.match(r'postgresql://([^:@]*):([^@]*)@([^/:]+)(?::(\d+))?/(.+)', DB_URL)
    if m:
        DB_CONFIG = {
            'user': m.group(1), 'password': m.group(2),
            'host': m.group(3), 'port': int(m.group(4) or 5432),
            'database': m.group(5).split('?')[0],
        }
    else:
        DB_CONFIG = {'database': 'xurshid_db'}

DATE = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT_DIR = '/var/backups/xurshid'
OUTPUT = f'{OUTPUT_DIR}/hisobot_{DATE}.xlsx'

# Stil konstantalar
HEADER_FILL = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
ALT_FILL   = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)


def style_header(ws, headers):
    ws.freeze_panes = 'A2'
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22


def add_rows(ws, rows, start=2):
    for r_idx, row in enumerate(rows, start):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 42)


def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. MAHSULOTLAR ────────────────────────────────────────────────
    ws = wb.create_sheet('📦 Mahsulotlar')
    headers = ['ID', 'Nomi', 'Barcode', 'Tan narx ($)', 'Sotish narx ($)',
               'Min zaxira', 'O\'lchov', 'Kategoriya']
    style_header(ws, headers)
    cur.execute("""
        SELECT p.id, p.name, COALESCE(p.barcode, '-'),
               ROUND(p.cost_price::numeric, 4),
               ROUND(p.sell_price::numeric, 4),
               p.min_stock, p.unit_type,
               COALESCE(c.name, '-')
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        ORDER BY p.name
    """)
    add_rows(ws, cur.fetchall())
    auto_width(ws)

    # ── 2. MIJOZLAR ───────────────────────────────────────────────────
    ws = wb.create_sheet('👥 Mijozlar')
    headers = ['ID', 'Ismi', 'Telefon', 'Qarz ($)', 'Balans ($)',
               'Oxirgi to\'lov', 'Qo\'shilgan sana']
    style_header(ws, headers)
    cur.execute("""
        SELECT c.id, c.name, COALESCE(c.phone, '-'),
               ROUND(COALESCE(
                   (SELECT SUM(debt_usd) FROM sales
                    WHERE customer_id = c.id AND payment_status = 'partial'), 0
               )::numeric, 2),
               ROUND(COALESCE(c.balance, 0)::numeric, 2),
               COALESCE(TO_CHAR(c.last_debt_payment_date, 'YYYY-MM-DD'), '-'),
               TO_CHAR(c.created_at, 'YYYY-MM-DD')
        FROM customers c
        ORDER BY c.name
    """)
    add_rows(ws, cur.fetchall())
    auto_width(ws)

    # ── 3. AKTIV QARZLAR ─────────────────────────────────────────────
    ws = wb.create_sheet('💰 Qarzlar')
    headers = ['Sana', 'Mijoz', 'Telefon', 'Qarz ($)', 'Qarz (UZS)',
               'To\'lov muddati', 'Sotuvchi']
    style_header(ws, headers)
    cur.execute("""
        SELECT TO_CHAR(s.sale_date, 'YYYY-MM-DD HH24:MI'),
               COALESCE(c.name, 'Noma''lum'),
               COALESCE(c.phone, '-'),
               ROUND(s.debt_usd::numeric, 2),
               ROUND(s.debt_amount::numeric, 0),
               COALESCE(TO_CHAR(s.payment_due_date, 'YYYY-MM-DD'), '-'),
               COALESCE(u.username, '-')
        FROM sales s
        LEFT JOIN customers c ON c.id = s.customer_id
        LEFT JOIN users u ON u.id = s.seller_id
        WHERE s.payment_status = 'partial' AND s.debt_usd > 0
        ORDER BY s.sale_date DESC
    """)
    rows = cur.fetchall()
    add_rows(ws, rows)
    # Jami qator
    if rows:
        cur.execute("SELECT ROUND(SUM(debt_usd)::numeric,2), ROUND(SUM(debt_amount)::numeric,0) FROM sales WHERE payment_status='partial' AND debt_usd > 0")
        totals = cur.fetchone()
        last = len(rows) + 2
        ws.cell(row=last, column=3, value='JAMI:').font = Font(bold=True)
        ws.cell(row=last, column=4, value=float(totals[0] or 0)).font = Font(bold=True)
        ws.cell(row=last, column=5, value=float(totals[1] or 0)).font = Font(bold=True)
    auto_width(ws)

    # ── 4. SOTUVLAR (so'nggi 30 kun) ──────────────────────────────────
    ws = wb.create_sheet('🛒 Sotuvlar (30 kun)')
    headers = ['Sana', 'Mijoz', 'Jami ($)', 'Foyda ($)',
               'Naqd ($)', 'Click ($)', 'Terminal ($)', 'Qarz ($)',
               'Holati', 'Sotuvchi']
    style_header(ws, headers)
    cur.execute("""
        SELECT TO_CHAR(s.sale_date, 'YYYY-MM-DD HH24:MI'),
               COALESCE(c.name, 'Noma''lum'),
               ROUND(s.total_amount::numeric, 2),
               ROUND(s.total_profit::numeric, 2),
               ROUND(s.cash_usd::numeric, 2),
               ROUND(s.click_usd::numeric, 2),
               ROUND(s.terminal_usd::numeric, 2),
               ROUND(s.debt_usd::numeric, 2),
               s.payment_status,
               COALESCE(u.username, '-')
        FROM sales s
        LEFT JOIN customers c ON c.id = s.customer_id
        LEFT JOIN users u ON u.id = s.seller_id
        WHERE s.sale_date >= NOW() - INTERVAL '30 days'
        ORDER BY s.sale_date DESC
    """)
    add_rows(ws, cur.fetchall())
    auto_width(ws)

    # ── 5. XARAJATLAR (so'nggi 30 kun) ───────────────────────────────
    ws = wb.create_sheet('💸 Xarajatlar (30 kun)')
    headers = ['Sana', 'Kategoriya', 'Tavsif',
               'Summa ($)', 'Summa (UZS)', 'Foydalanuvchi']
    style_header(ws, headers)
    cur.execute("""
        SELECT TO_CHAR(e.expense_date, 'YYYY-MM-DD HH24:MI'),
               COALESCE(e.category, '-'),
               COALESCE(e.description, COALESCE(e.title, '-')),
               ROUND(COALESCE(e.amount_usd, 0)::numeric, 2),
               ROUND(COALESCE(e.amount_uzs, 0)::numeric, 0),
               COALESCE(e.created_by, '-')
        FROM expenses e
        WHERE e.expense_date >= NOW() - INTERVAL '30 days'
        ORDER BY e.expense_date DESC
    """)
    add_rows(ws, cur.fetchall())
    auto_width(ws)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT)
    conn.close()
    print(OUTPUT)


if __name__ == '__main__':
    main()
