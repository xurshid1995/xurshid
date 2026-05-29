# -*- coding: utf-8 -*-
"""Ustun sarlavhalari va ma'lumotlari mosligini tekshirish"""
import os, sys, glob
sys.path.insert(0, '/var/www/xurshid')
from dotenv import load_dotenv
load_dotenv('/var/www/xurshid/.env')
import openpyxl

files = sorted(glob.glob('/var/backups/xurshid/hisobot_*.xlsx'))
if not files:
    print("Excel fayl topilmadi!")
    sys.exit(1)

latest = files[-1]
wb = openpyxl.load_workbook(latest)

for ws in wb.worksheets:
    print()
    print("=" * 70)
    print(f"  {ws.title}  ({ws.max_row-1} qator)")
    print("=" * 70)
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    # 3 ta misol qator (header=1, data=2,3,4)
    show_rows = min(ws.max_row, 4)
    for r in range(2, show_rows + 1):
        print(f"\n  --- {r-1}-qator ---")
        for c_idx, header in enumerate(headers, 1):
            val = ws.cell(r, c_idx).value
            print(f"    [{c_idx:2}] {str(header):<25} => {repr(val)}")
    
    # Oxirgi qatorni ham ko'rsatish (jami qator bo'lsa)
    if ws.max_row > 4:
        last_r = ws.max_row
        print(f"\n  --- OXIRGI qator ({last_r-1}-qator) ---")
        for c_idx, header in enumerate(headers, 1):
            val = ws.cell(last_r, c_idx).value
            if val is not None:
                print(f"    [{c_idx:2}] {str(header):<25} => {repr(val)}")
